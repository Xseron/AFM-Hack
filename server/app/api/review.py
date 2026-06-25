from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import Response

from app.api.deps import get_components
from app.api.serializers import job_list_item
from app.pipelines import aggregator

router = APIRouter()


@router.get("/review-queue")
async def review_queue(limit: int = 50, components=Depends(get_components)) -> dict:
    jobs = await components.repo.review_queue(limit=limit)
    return {
        "items": [job_list_item(j) for j in jobs]
    }


@router.get("/priority-list")
async def priority_list(limit: int = 50, components=Depends(get_components)) -> dict:
    jobs = await components.repo.priority_queue(limit=limit)
    return {
        "items": [job_list_item(j) for j in jobs]
    }


@router.get("/recent-jobs")
async def recent_jobs(limit: int = 50, components=Depends(get_components)) -> dict:
    jobs = await components.repo.recent_jobs(limit=limit)
    return {
        "items": [job_list_item(j) for j in jobs]
    }


# Scanners that do not drive the manual-check risk level (kept in sync with the
# dashboard's RISK_EXCLUDE set in app.api.ui).
RISK_EXCLUDE = {"contact_spam", "deepfake_gend"}


def _risk_summary(scores: dict[str, float]) -> tuple[str, str | None, float]:
    """(risk level, top scanner, top score) over non-excluded scanners.

    Mirrors the dashboard rule: red if any non-excluded scanner exceeds its
    threshold, yellow if one reaches threshold/1.5, green otherwise.
    """
    level = 0  # 0 green, 1 yellow, 2 red
    top_name: str | None = None
    top_val = 0.0
    for name, raw in (scores or {}).items():
        if name in RISK_EXCLUDE:
            continue
        value = float(raw)
        if top_name is None or value > top_val:
            top_name, top_val = name, value
        threshold = aggregator.threshold_for(name)
        if value > threshold:
            level = max(level, 2)
        elif value >= threshold / 1.5:
            level = max(level, 1)
    return ({0: "green", 1: "yellow", 2: "red"}[level], top_name, top_val)


@router.get("/export.xlsx")
async def export_xlsx(limit: int = 10000, components=Depends(get_components)) -> Response:
    """Download all jobs as an .xlsx, sorted high->low manual-check priority.

    Rows are shaded red/yellow/green by the same rule the dashboard uses, and the
    "Top signal"/score columns surface the highest-scoring scanner (the likely
    scam category) for each reel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    jobs = await components.repo.recent_jobs(limit=limit)
    items = [job_list_item(j) for j in jobs]

    # Union of scanner ids across all jobs -> one confidence column each.
    scanner_ids: list[str] = []
    seen: set[str] = set()
    for item in items:
        for name in (item.get("scanner_confidences") or {}):
            if name not in seen:
                seen.add(name)
                scanner_ids.append(name)
    scanner_ids.sort()

    rows = []
    for item in items:
        scores = item.get("scanner_confidences") or {}
        level, top_name, top_val = _risk_summary(scores)
        rows.append((item, scores, level, top_name, top_val))
    # Manual-check priority: highest non-excluded confidence first.
    rows.sort(key=lambda r: r[4], reverse=True)

    headers = [
        "Rank", "Risk", "Top score", "Top signal", "Category", "Risk score",
        "Priority", "Status", "Platform", "Shortcode", "Link", "Created",
        "Job ID", "Description",
    ] + [f"score: {name}" for name in scanner_ids]

    wb = Workbook()
    ws = wb.active
    ws.title = "Manual review"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1769E0")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    fills = {
        "red": PatternFill("solid", fgColor="F8D7DA"),
        "yellow": PatternFill("solid", fgColor="FFF3CD"),
        "green": PatternFill("solid", fgColor="D4EDDA"),
    }

    for rank, (item, scores, level, top_name, top_val) in enumerate(rows, start=1):
        source = item.get("source") or {}
        link = source.get("permalink") or source.get("top_bar_url") or source.get("url") or ""
        row = [
            rank,
            level.upper(),
            round(float(top_val), 4),
            top_name or "",
            item.get("category") or "",
            item.get("risk_score") if isinstance(item.get("risk_score"), (int, float)) else "",
            item.get("priority") if isinstance(item.get("priority"), (int, float)) else "",
            item.get("status") or "",
            source.get("platform") or "",
            source.get("shortcode") or "",
            link,
            item.get("created_at") or "",
            item.get("job_id") or "",
            (item.get("description") or "")[:500],
        ]
        row += [
            round(float(scores[name]), 4) if name in scores else ""
            for name in scanner_ids
        ]
        ws.append(row)
        excel_row = rank + 1
        for cell in ws[excel_row]:
            cell.fill = fills[level]
        link_cell = ws.cell(row=excel_row, column=headers.index("Link") + 1)
        if link:
            link_cell.hyperlink = link
            link_cell.font = Font(color="1769E0", underline="single")

    ws.freeze_panes = "A2"
    widths = {
        "A": 6, "B": 8, "C": 10, "D": 16, "E": 12, "F": 10, "G": 9, "H": 10,
        "I": 12, "J": 16, "K": 40, "L": 22, "M": 24, "N": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="media_watch_review_{stamp}.xlsx"'},
    )
